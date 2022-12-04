from tkinter import *
import random
import openpyxl

is_run = False

def getPeopleList():
    workbook = openpyxl.load_workbook("./peopleList.xlsx", data_only=True)
    list = []
    sheet = workbook[workbook.sheetnames[0]]
    for row in range(2,sheet.max_row+1):
        if sheet.cell(row=row,column=2).value == "":
            continue
        else:
            list.append(sheet.cell(row=row,column=2).value)
    return list

def randomRun(list):
    global is_run
    if is_run:#此部分用于避免多次点击开始造成程序重复执行的错误
        return
    is_run = True
    num = insert_point()
    if num:
        num = int(num)
        randomResult(list,num)
    else:
        is_run = False

def insert_point():
     var = input.get() #获取输入的信息
     return var

def randomResult(list,num):
    global is_run
    result = random.sample(list, num)
    var.set(result)
    if is_run:
        window.after(100, randomResult, list, num)

def finalResult():
    global is_run
    is_run = False

if __name__ == '__main__': #主程序
    window = Tk() #初始化TK窗口
    window.geometry('500x290+250+150') #设定窗口大小
    window.resizable(0,0) #将窗口设置为不可拉伸
    window.title('抽奖小程序')
    list = getPeopleList() #从peopleList.xlsx中获取待抽奖人员名单

    var = StringVar() #初始化一个字符串变量，用于滚动显示抽奖结果

    noteLable = Label(text="请输入抽奖人数：")
    noteLable.place(anchor=NW, x=120, y=30)
    input = Entry(window, show=None) #设置一个输入框，用于输入抽奖人数
    input.place(anchor=NW, x=240, y=30)

    resultLable = Label(textvariable=var) #设置一个显示抽奖结果的文本框
    resultLable.place(anchor=NW,x=150,y=100)

    startBt = Button(text="开始", command=lambda: randomRun(list=list)) #开始按钮
    confirmBt = Button(text="确定", command=lambda: finalResult()) #停止按钮
    startBt.place(anchor=NW, x=200, y=180)
    confirmBt.place(anchor=NW, x=260, y=180)

    window.mainloop() #渲染窗口