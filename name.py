import os

from openpyxl import load_workbook
wb = load_workbook('C:\\Users\\admin\\Desktop\\file_name.xlsx')
sheet = wb['Sheet1']
dic = {}
for row_index in range(1, sheet.max_row + 1):
    temp = sheet.cell(row=row_index, column=2).value
    oldName = str(sheet.cell(row=row_index, column=1).value) + ".docx"
    newName = temp + ".docx"
    dic[oldName] = newName
path = "C:\\Users\\admin\\Desktop\\作业"
files = os.listdir(path)
for fileName in files:
    if fileName.endswith(".docx"):
        os.rename(path + os.sep + fileName, path + os.sep + dic[fileName])
